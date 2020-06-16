import csv
import logging
from typing import Union, Optional
from decimal import Decimal
from collections import namedtuple

from openpyxl import load_workbook

from server.pm.models import Component, Collection, Manufacturer, ComponentImport

logger = logging.getLogger(__name__)

Column = namedtuple('Column', ['field', 'columns', 'is_required'])

COLUMNS = (
    Column(field='code', columns={'code'}, is_required=True),
    Column(field='name', columns={'name'}, is_required=True),
    Column(field='description', columns={'description'}, is_required=False),
    Column(field='collection', columns={'collection'}, is_required=False),
    Column(field='manufacturer', columns={'manufacturer'}, is_required=False),
    Column(field='price', columns={'price'}, is_required=True),
)

REQUIRED = {col.field for col in COLUMNS if col.is_required}
ALL = {col.field for col in COLUMNS}


def process_component(
    code: str,
    name: str,
    price: Union[str, int, float, Decimal],
    description: Optional[str] = None,
    collection: Optional[str] = None,
    manufacturer: Optional[str] = None,
) -> None:
    code, name = code.strip(), name.strip()
    description = description and description.stip()
    collection = collection and collection.strip()
    manufacturer = manufacturer and manufacturer.strip()

    collection_instance = None
    if collection:
        collection_instance, _ = Collection.objects.get_or_create(name=collection)

    manufacturer_instance = None
    if manufacturer:
        manufacturer_instance, _ = Manufacturer.objects.get_or_create(name=manufacturer)

    component = Component.objects.filter(code=code, collection=collection_instance)
    if component.exists():
        component = component.get()
        component.name = name
        component.description = description or component.description
        component.price = price
        component.manufacturer = manufacturer_instance or component.manufacturer
        component.save(update_fields=['name', 'description', 'price', 'manufacturer'])
    else:
        Component.objects.create(
            code=code,
            name=name,
            price=price,
            description=description,
            collection=collection_instance,
            manufacturer=manufacturer_instance,
        )


def get_header_mapping(row):
    mapping = {}
    for index, value in enumerate(row):
        strval = str(value or '').lower()
        for col in COLUMNS:
            if strval in col.columns:
                mapping[col.field] = index
                break

    return mapping if mapping and set(mapping.keys()).issuperset(REQUIRED) else {}


def process_rows(iter_rows, name):
    rows = 0
    for row in iter_rows:
        rows += 1
        mapping = get_header_mapping(row)
        if mapping:
            break
    else:
        return 0, 0, [{'type': 'error', 'sheet': name, 'error': 'Missing header row'}]

    processed = 0
    errors = []
    for row in iter_rows:
        rows += 1
        component = {field: row[index] for field, index in mapping.items()}
        component['code'] = str(component['code'])
        fields = {key for key, value in component.items() if value}
        if fields.issuperset(REQUIRED):
            process_component(**component)
            processed += 1
        elif fields.intersection(ALL):
            errors.append(
                {
                    'type': 'warning',
                    'sheet': name,
                    'row': rows,
                    'error': 'Missing required fields {0}'.format(', '.join(fields.difference(REQUIRED))),
                }
            )

    return rows, processed, errors


def import_from_workbook(import_info: ComponentImport):
    wb = load_workbook(import_info.full_path)

    for sheet_name in wb.sheetnames:
        try:
            rows, processed, sheet_errors = process_rows(wb[sheet_name].values, sheet_name)
        except Exception as exc:
            rows, processed = 0, 0
            sheet_errors = [{'type': 'error', 'sheet': sheet_name, 'error': str(exc)}]
        finally:
            import_info.rows = (import_info.rows or 0) + rows
            import_info.processed = (import_info.processed or 0) + processed
            import_info.errors.extend(sheet_errors)
            import_info.save(update_fields=['rows', 'processed', 'errors'])


def import_csv(import_info: ComponentImport):
    with open(import_info.full_path, 'r') as f:
        reader = csv.reader(f)

        try:
            rows, processed, sheet_errors = process_rows(reader, 'csv')
        except Exception as exc:
            rows, processed = 0, 0
            sheet_errors = [{'type': 'error', 'sheet': 'csv', 'error': str(exc)}]
        finally:
            import_info.rows = (import_info.rows or 0) + rows
            import_info.processed = (import_info.processed or 0) + processed
            import_info.errors.extend(sheet_errors)
            import_info.save(update_fields=['rows', 'processed', 'errors'])


def import_components(import_uuid: str):
    try:
        import_info = ComponentImport.objects.get(uuid=import_uuid)
    except ComponentImport.DoesNotExist:
        logger.error('Missing component import task, uuid=%s', import_uuid)
        return

    if import_info.status != ComponentImport.ImportStatus.QUEUED.value:
        logger.error('Got import task with invalid status, status=%s, uuid=%s', import_info.status, import_uuid)
        return

    import_info.status = ComponentImport.ImportStatus.PROCESSING.value
    import_info.save(update_fields=['status'])

    extension = import_info.import_file.split('.')[-1]
    if extension == 'xlsx':
        import_from_workbook(import_info)
        import_info.status = ComponentImport.ImportStatus.COMPLETE.value
    elif extension == 'csv':
        import_csv(import_info)
        import_info.status = ComponentImport.ImportStatus.COMPLETE.value
    else:
        import_info.status = ComponentImport.ImportStatus.FAILED.value

    import_info.save(update_fields=['status'])
